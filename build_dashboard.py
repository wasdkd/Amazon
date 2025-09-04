import datetime, glob, json, os, textwrap
from openpyxl import load_workbook

SRC_DIR = "uploads"  # GitHub 上的 excel 目录
OUT_FILE = "ExcelAPI.html"  # 输出到仓库根目录
API_DIR = "api"  # 静态 js 目录（不动）


# -------------------- 原逻辑函数 --------------------
def color_block(r, c, h_blocks, v_blocks):
    h_key = h_blocks.get(c, c)
    v_key = v_blocks.get(r, r)
    palette = ["#fce4ec", "#e3f2fd", "#ddfbb8", "#e8f5e9", "#fce8e6", "#f3e5f5", "#b9ffe8", "#fff8e1"]
    return palette[(h_key + v_key) % len(palette)]


def build_blocks(ws):
    h_blocks, v_blocks = {}, {}
    for m in ws.merged_cells.ranges:
        if m.min_row == 1:
            for c in range(m.min_col, m.max_col + 1):
                h_blocks[c] = m.min_col
        if m.min_col == 2:
            for r in range(m.min_row, m.max_row + 1):
                v_blocks[r] = m.min_row
    return h_blocks, v_blocks


def excel_to_html(excel_path, sheet_index=0):
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.worksheets[sheet_index]
    rows = list(ws.iter_rows())

    h_blocks, v_blocks = build_blocks(ws)
    row_height = 38

    # 合并映射
    merge_map = {
        (m.min_row, m.min_col, m.max_row, m.max_col): str(ws.cell(m.min_row, m.min_col).value or "")
        for m in ws.merged_cells.ranges
    }

    html_lines = ['<table class="main-table"><thead>']
    # 表头两行
    for r in (1, 2):
        html_lines.append("<tr>")
        for c, cell in enumerate(rows[r - 1], 1):
            val = cell.value or ""
            sty = f"background:#c8e6c9;color:#000;height:{row_height}px"
            merged = None
            for (min_r, min_c, max_r, max_c), v in merge_map.items():
                if (r, c) == (min_r, min_c):
                    rowspan = max_r - min_r + 1
                    colspan = max_c - min_c + 1
                    html_lines.append(
                        f'<th rowspan="{rowspan}" colspan="{colspan}" '
                        f'style="{sty};position:sticky;top:{(r - 1) * row_height}px;left:0;z-index:4;">{v}</th>'
                    )
                    merged = True
                    break
                elif min_r <= r <= max_r and min_c <= c <= max_c:
                    merged = True
                    break
            if not merged:
                html_lines.append(
                    f'<th style="{sty};position:sticky;top:{(r - 1) * row_height}px;left:0;z-index:4;">{val}</th>' if c == 1 else
                    f'<th style="{sty};position:sticky;top:{(r - 1) * row_height}px;z-index:3;">{val}</th>'
                )
        html_lines.append("</tr>")
    html_lines.append("</thead><tbody>")

    # 数据区
    for r, row in enumerate(rows[2:], 3):
        html_lines.append("<tr>")
        for c, cell in enumerate(row, 1):
            val = cell.value or ""
            bg = color_block(r, c, h_blocks, v_blocks)
            merged = None
            for (min_r, min_c, max_r, max_c), v in merge_map.items():
                if (r, c) == (min_r, min_c):
                    rowspan = max_r - min_r + 1
                    colspan = max_c - min_c + 1
                    html_lines.append(
                        f'<td rowspan="{rowspan}" colspan="{colspan}" '
                        f'style="background:{bg};color:#000;position:sticky;left:0;z-index:2;">{v}</td>'
                    )
                    merged = True
                    break
                elif min_r <= r <= max_r and min_c <= c <= max_c:
                    merged = True
                    break
            if not merged:
                html_lines.append(
                    f'<td style="background:{bg};color:#000;position:sticky;left:0;z-index:2;">{val}</td>' if c == 1 else
                    f'<td style="background:{bg};color:#000;">{val}</td>'
                )
        html_lines.append("</tr>")
    html_lines.append("</tbody></table>")
    return "\n".join(html_lines)


# -------------------- 生成页面 --------------------
def build_dashboard():
    files = glob.glob(os.path.join(SRC_DIR, "*.xlsx"))
    table_data = {}
    cards_html = ""
    for f in files:
        fname = os.path.basename(f)
        key = "".join(c for c in fname if c.isalnum() or c in (' ', '.', '_')).rstrip().replace(" ", "_").replace(".",
                                                                                                                  "_")
        table_data[key] = excel_to_html(f)
        mtime = datetime.datetime.fromtimestamp(os.path.getmtime(f)).strftime('%Y-%m-%d %H:%M')
        size = os.path.getsize(f)
        size_str = f"{size // 1024}KB" if size < 1024 * 1024 else f"{size // (1024 * 1024):.1f}MB"
        cards_html += f"""
        <div class="card">
            <div class="card-header">{fname}</div>
            <div class="card-body">
                <div class="file-info">
                    <span>大小: {size_str}</span>
                    <span>修改时间: {mtime}</span>
                </div>
                <div class="view-button" onclick="showTable('{key}', {repr(fname)})">查看表格数据</div>
                <!-- 删除/覆盖按钮 -->
                <div style="margin-top:0.5rem;display:flex;gap:0.5rem;">
                  <button class="view-button" style="background:#dc3545;font-size:0.75rem;" onclick="deleteFile({repr(fname)})">🗑 删除</button>
                  <button class="view-button" style="background:#ffc107;font-size:0.75rem;" onclick="overrideFile({repr(fname)})">🔄 覆盖</button>
                </div>
            </div>
        </div>
        """

    css = textwrap.dedent(
        """


    :root {
                    --primary: #0d6efd;
                    --primary-dark: #0b5ed7;
                    --secondary: #6c757d;
                    --success: #198754;
                    --light: #f8f9fa;
                    --dark: #212529;
                    --gray-100: #f8f9fa;
                    --gray-200: #e9ecef;
                    --gray-300: #dee2e6;
                    --gray-400: #ced4da;
                    --gray-500: #adb5bd;
                    --gray-800: #343a40;
                    --border-radius: 8px;
                    --box-shadow: 0 0.125rem 0.25rem rgba(0, 0, 0, 0.075);
                    --box-shadow-hover: 0 0.5rem 1rem rgba(0, 0, 0, 0.15);
                    --transition: all 0.2s ease-in-out;
                }

                * {
                    margin: 0;
                    padding: 0;
                    box-sizing: border-box;
                }

                body {
                    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
                    font-size: 16px;
                    background-color: #f8f9fa;
                    color: var(--gray-800);
                    line-height: 1.5;
                }

                .container {
                    width: 100%;
                    max-width: 100%;
                    padding: 0;
                    margin: 0;
                }

                .header {
                    background: linear-gradient(120deg, var(--primary), #0b5ed7);
                    color: white;
                    padding: 1rem 2rem;
                    box-shadow: var(--box-shadow);
                    display: flex;
                    justify-content: space-between;
                    align-items: center;
                    flex-wrap: wrap;
                }

                .header h1 {
                    font-size: 1.5rem;
                    font-weight: 600;
                }

                .header-time {
                    font-size: 1rem;
                    opacity: 0.95;
                }

                .search-container {
                    display: flex;
                    gap: 0.5rem;
                    align-items: center;
                    margin: 0 1rem;
                }

                .search-input {
                    padding: 0.5rem;
                    border-radius: var(--border-radius);
                    border: 1px solid var(--gray-300);
                    font-size: 1rem;
                    min-width: 200px;
                }

                .search-button {
                    background: var(--success);
                    color: white;
                    border: none;
                    padding: 0.5rem 1rem;
                    border-radius: var(--border-radius);
                    cursor: pointer;
                    font-weight: 500;
                }

                .search-button:hover {
                    background: #157347;
                }

                .nav-buttons {
                    display: flex;
                    gap: 0.5rem;
                    align-items: center;
                }

                .nav-button {
                    background: #6f42c1;
                    color: white;
                    border: none;
                    padding: 0.5rem;
                    border-radius: var(--border-radius);
                    cursor: pointer;
                    font-weight: 500;
                }

                .nav-button:hover {
                    background: #5a32a3;
                }

                .nav-button:disabled {
                    background: var(--gray-500);
                    cursor: not-allowed;
                }

                .search-info {
                    color: white;
                    font-size: 0.9rem;
                    min-width: 120px;
                    text-align: center;
                }

                .grid {
                    display: grid;
                    grid-template-columns: repeat(auto-fit, minmax(350px, 1fr));
                    gap: 1.5rem;
                    padding: 1.5rem;
                    justify-content: center;
                    max-width: 1500px;
                    margin: 0 auto;
                }

                .grid-wrapper {
                    display: flex;
                    justify-content: center;
                    width: 100%;
                }

                .card {
                    background: white;
                    border: 1px solid var(--gray-300);
                    border-radius: var(--border-radius);
                    overflow: hidden;
                    box-shadow: var(--box-shadow);
                    transition: var(--transition);
                    margin: 0 auto;
                    max-width: 350px;
                }

                .card:hover {
                    transform: translateY(-0.25rem);
                    box-shadow: var(--box-shadow-hover);
                }

                .card-header {
                    padding: 0.75rem 1rem;
                    background: linear-gradient(120deg, var(--primary), #0b5ed7); /* 修改为更鲜明的颜色 */
                    color: white; /* 添加白色文字 */
                    border-bottom: 1px solid var(--gray-300);
                    font-weight: 600;
                    font-size: 1.1rem;
                }

                .card-body {
                    padding: 1.25rem;
                }

                .file-info {
                    display: flex;
                    justify-content: space-between;
                    margin-bottom: 1.25rem;
                    font-size: 0.875rem;
                    color: var(--secondary);
                }

                .file-info span {
                    background-color: var(--gray-100);
                    padding: 0.25rem 0.75rem;
                    border-radius: 20px;
                }

                .view-button {
                    display: inline-block;
                    background: var(--primary);
                    color: white;
                    padding: 0.5rem 1rem;
                    border: none;
                    border-radius: var(--border-radius);
                    text-decoration: none;
                    font-weight: 500;
                    transition: var(--transition);
                    width: 100%;
                    text-align: center;
                    box-shadow: 0 2px 4px rgba(0, 0, 0, 0.1);
                    cursor: pointer;
                }

                .view-button:hover {
                    background: var(--primary-dark);
                    transform: scale(1.02);
                    box-shadow: 0 4px 8px rgba(0, 0, 0, 0.15);
                }

                .back-button {
                    background: #40a2ff;
                    color: white;
                    padding: 0.5rem 1rem;
                    border: none;
                    border-radius: var(--border-radius);
                    text-decoration: none;
                    font-weight: 500;
                    transition: var(--transition);
                    display: inline-flex;
                    align-items: center;
                    gap: 0.5rem;
                    box-shadow: 0 2px 4px rgba(0, 0, 0, 0.1);
                    cursor: pointer;
                }

                .back-button:hover {
                    background: var(--primary-dark);
                    transform: scale(1.02);
                    box-shadow: 0 4px 8px rgba(0, 0, 0, 0.15);
                }

                .table-container {
                    padding: 1rem;
                    display: none;
                }

                .table-wrapper {
                    height: 85vh;
                    overflow: auto;
                    border-top: 1px solid var(--gray-300);
                }

                .main-table {
                    border-collapse: collapse;
                    width: 100%;
                }

                .main-table th,
                .main-table td {
                    border: 1px solid var(--gray-400);
                    padding: 6px 8px;
                    text-align: center;
                    white-space: nowrap;
                }

                .main-table th {
                    position: sticky;
                    top: 0;
                    z-index: 3;
                    background: #c8e6c9;
                    color: #000;
                }

                .main-table td:first-child,
                .main-table th:first-child {
                    position: sticky;
                    left: 0;
                    z-index: 2;
                    background: inherit;
                }

                /* 搜索高亮样式 */
                .highlight {
                    background-color: #ffeb3b !important;
                    color: #000 !important;
                    font-weight: bold;
                }

                .current-highlight {
                    background-color: #ff9800 !important;
                    color: #fff !important;
                    font-weight: bold;
                }

                /* 列高亮样式 */
                .column-highlight {
                    background-color: rgba(173, 216, 230, 0.5) !important;
                }

                /* 十字高亮样式 */
                .cross-highlight-row {
                    background-color: rgba(255, 255, 0, 0.2) !important;
                }

                .cross-highlight-col {
                    background-color: rgba(255, 255, 0, 0.2) !important;
                }

                .cross-highlight-cell {
                    background-color: rgba(255, 255, 0, 0.3) !important;
                    position: relative;
                }

                .cross-highlight-cell::after {
                    content: "";
                    position: absolute;
                    top: 0;
                    left: 0;
                    right: 0;
                    bottom: 0;
                    border: 2px solid #ff0;
                    pointer-events: none;
                }

                .footer {
                    text-align: center;
                    padding: 2px;
                    color: var(--secondary);
                    font-size: 0.875rem;
                    border-top: 1px solid var(--gray-300);
                    margin-top: 1px;
                }

                .file-title {
                    display: flex;
                    justify-content: space-between;
                    align-items: center;
                    padding: 5px 20px;
                    background: white;
                    border-bottom: 1px solid var(--gray-300);
                    flex-wrap: wrap;
                    gap: 1rem;
                }

                .file-name {
                    font-size: 1.25rem;
                    font-weight: 600;
                    color: var(--gray-800);
                }

                /* 隐藏的表格容器 */
                .hidden-table {
                    display: none;
                }

                /* 显示的表格容器 */
                .visible-table {
                    display: block;
                }

                @media (max-width: 768px) {
                    .grid {
                        grid-template-columns: 1fr;
                        padding: 1rem;
                    }

                    .header {
                        flex-direction: column;
                        text-align: center;
                        gap: 1rem;
                        padding: 1rem;
                    }

                    .file-title {
                        padding: 1rem;
                    }

                    .search-container {
                        width: 100%;
                        margin: 0.5rem 0;
                        justify-content: center;
                    }
                }
    """)

    js = textwrap.dedent(f"""
        <script>
        window.GH = {{
          OWNER: "wasdkd",
          REPO : "Amazon",
          BRANCH: "main",
          getToken() {{
            return localStorage.getItem('github_token');
          }},
          setToken(token) {{
            localStorage.setItem('github_token', token);
          }},
          clearToken() {{
            localStorage.removeItem('github_token');
          }},
          async callGitHub(method, path, body=null, token=""){{
            const url = `https://api.github.com/repos/${{this.OWNER}}/${{this.REPO}}/contents/${{path}}`;
            const opt = {{ method, headers: {{
              "Authorization": `token ${{token}}`,
              "Accept": "application/vnd.github.v3+json",
              "Content-Type": "application/json"
            }}}};
            if(body) opt.body = JSON.stringify(body);
            const r = await fetch(url, opt);
            if(!r.ok) throw new Error(await r.text());
            return r.json();
          }}
        }};
        window.uploadFile = async function (){{
          let tok = GH.getToken();
          if (!tok) {{
            tok = prompt("请输入 GitHub Personal Access Token：");
            if (!tok) return;
            GH.setToken(tok);
          }}
          const inp = document.createElement("input");
          inp.type = "file"; inp.accept = ".xlsx";
          inp.onchange = async (e) => {{
            const file = e.target.files[0];
            if(!file) return;
            const content = await file.arrayBuffer().then(b => btoa(String.fromCharCode(...new Uint8Array(b))));
            const path = `uploads/${{file.name}}`;
            try{{ 
              await GH.callGitHub("PUT", path, {{message:`upload ${{file.name}}`, content, branch:GH.BRANCH}}, tok);
              alert("上传成功！即将刷新"); location.reload();
            }}catch(err){{ 
              if (err.message.includes('Bad credentials')) {{
                alert("Token无效，请重新输入");
                GH.clearToken();
              }} else {{
                alert("失败："+err); 
              }}
            }}
          }}; inp.click();
        }};
        window.deleteFile = async function (fileName) {{
          let tok = GH.getToken();
          if (!tok) {{
            tok = prompt("请输入 GitHub Personal Access Token：");
            if (!tok) return;
            GH.setToken(tok);
          }}
          const path = `uploads/${{fileName}}`;
          try{{
            const {{sha}} = await GH.callGitHub("GET", path, null, tok);
            await GH.callGitHub("DELETE", path, {{message:`delete ${{fileName}}`, sha, branch:GH.BRANCH}}, tok);
            alert("删除成功！即将刷新"); location.reload();
          }}catch(err){{ 
            if (err.message.includes('Bad credentials')) {{
              alert("Token无效，请重新输入");
              GH.clearToken();
            }} else {{
              alert("删除失败：" + err); 
            }}
          }}
        }};
        window.overrideFile = function (fileName) {{
          let tok = GH.getToken();
          if (!tok) {{
            tok = prompt("请输入 GitHub Personal Access Token：");
            if (!tok) return;
            GH.setToken(tok);
          }}
          const inp = document.createElement("input");
          inp.type = "file"; inp.accept = ".xlsx";
          inp.onchange = async (e) => {{
            const file = e.target.files[0];
            if(!file) return;
            const content = await file.arrayBuffer().then(b => btoa(String.fromCharCode(...new Uint8Array(b))));
            const path = `uploads/${{fileName}}`;
            try{{ 
              const {{sha}} = await GH.callGitHub("GET", path, null, tok);
              await GH.callGitHub("DELETE", path, {{message:`override ${{fileName}}`, sha, branch:GH.BRANCH}}, tok);
              await GH.callGitHub("PUT", path, {{message:`override ${{fileName}}`, content, branch:GH.BRANCH}}, tok);
              alert("覆盖成功！即将刷新"); location.reload();
            }}catch(err){{ 
              if (err.message.includes('Bad credentials')) {{
                alert("Token无效，请重新输入");
                GH.clearToken();
              }} else {{
                alert("覆盖失败：" + err); 
              }}
            }}
          }}; inp.click();
        }};
        window.clearSavedToken = function() {{
          GH.clearToken();
          alert("已清除保存的Token");
        }};
        const tableData = {json.dumps(table_data)};
        let searchResults = [], currentResultIndex = -1, searchTerm = '';
        let currentHighlightedCell = null, highlightedRows = [], highlightedCols = [];

        function showTable(fileKey, fileName) {{
          document.querySelector('.grid-wrapper').style.display = 'none';
          document.querySelector('.header h1').textContent = fileName;
          document.getElementById('backButton').style.display = 'inline-flex';
          const container = document.getElementById('tableContainer');
          container.innerHTML = '<div class="table-wrapper">' + tableData[fileKey] + '</div>';
          container.style.display = 'block';
          document.getElementById('searchControls').style.display = 'flex';
          if (searchTerm) performSearch(false);
          setTimeout(() => {{ const t = document.querySelector('.main-table'); if (t) t.addEventListener('click', handleCellClick); }}, 100);
        }}
        function goBack() {{
          document.querySelector('.grid-wrapper').style.display = 'flex';
          document.querySelector('.header h1').textContent = '企划数据看板';
          document.getElementById('backButton').style.display = 'none';
          document.getElementById('tableContainer').style.display = 'none';
          document.getElementById('searchControls').style.display = 'none';
          clearSearch(); clearCrossHighlight();
        }}
        function handleCellClick(e) {{
      const cell = e.target; if (cell.tagName !== 'TD' && cell.tagName !== 'TH') return;
      clearCrossHighlight(); currentHighlightedCell = cell; cell.classList.add('cross-highlight-cell');
      const table = cell.closest('table'), allRows = Array.from(table.querySelectorAll('tr'));
      const cellRect = cell.getBoundingClientRect(), cellCenterX = (cellRect.left + cellRect.right)/2, cellCenterY = (cellRect.top + cellRect.bottom)/2;
      allRows.forEach(row => {{ 
        const cells = Array.from(row.querySelectorAll('td, th')); 
        cells.forEach(targetCell => {{
          const tRect = targetCell.getBoundingClientRect(), tCenterX = (tRect.left + tRect.right)/2, tCenterY = (tRect.top + tRect.bottom)/2;
          if (Math.abs(tCenterY - cellCenterY) < 2) {{ 
            targetCell.classList.add('cross-highlight-row'); 
            highlightedRows.push(targetCell); 
          }}
          if (Math.abs(tCenterX - cellCenterX) < 2) {{ 
            targetCell.classList.add('cross-highlight-col'); 
            highlightedCols.push(targetCell); 
          }}
        }}); 
      }}); 
    }}
        function clearCrossHighlight() {{
          if (currentHighlightedCell) {{ currentHighlightedCell.classList.remove('cross-highlight-cell'); currentHighlightedCell = null; }}
          highlightedRows.forEach(c => c.classList.remove('cross-highlight-row')); highlightedRows = [];
          highlightedCols.forEach(c => c.classList.remove('cross-highlight-col')); highlightedCols = [];
        }}
        function performSearch(scroll=true) {{
          const searchInput = document.getElementById('searchInput');
          const newSearchTerm = searchInput ? searchInput.value.trim() : '';
          if (!newSearchTerm) {{ alert('请输入搜索关键词'); return; }}
          const tableWrapper = document.querySelector('.table-wrapper');
          if (!tableWrapper) {{ alert('请先选择一个表格文件'); return; }}
          if (searchTerm !== newSearchTerm || searchResults.length === 0) {{
            searchTerm = newSearchTerm; clearHighlights();
            const regex = new RegExp('(' + searchTerm.replace(/[.*+?^${{}}()|[\\]\\\\]/g, '\\\\$&') + ')', 'gi');
            const tableCells = tableWrapper.querySelectorAll('td, th');
            tableCells.forEach((cell, idx) => {{ if (regex.test(cell.textContent)) {{ searchResults.push({{ element: cell, index: idx }}); }} }});
            if (searchResults.length === 0) {{ alert('未找到匹配项'); updateSearchInfo(); return; }}
            searchResults.forEach((res, i) => {{ res.element.innerHTML = res.element.innerHTML.replace(regex, '<span class="highlight" data-highlight-id="hl-${{i}}">$1</span>'); }});
          }}
          if (scroll) {{ currentResultIndex = -1; nextResult(); }} else if (currentResultIndex >= 0 && currentResultIndex < searchResults.length) {{ updateCurrentHighlight(); }}
          updateSearchInfo();
        }}
        function updateCurrentHighlight() {{
          document.querySelectorAll('.current-highlight').forEach(el => el.classList.remove('current-highlight'));
          if (currentResultIndex >= 0 && currentResultIndex < searchResults.length) {{ const h = searchResults[currentResultIndex].element.querySelector('.highlight'); if (h) h.classList.add('current-highlight'); }}
        }}
        function clearHighlights() {{
          document.querySelectorAll('.highlight').forEach(span => {{ span.outerHTML = span.innerHTML; }});
          searchResults = []; currentResultIndex = -1;
        }}
        function clearSearch() {{
          const searchInput = document.getElementById('searchInput'); if (searchInput) searchInput.value = '';
          searchTerm = ''; clearHighlights(); updateSearchInfo();
        }}
        function updateSearchInfo() {{
          const info = document.getElementById('searchInfo');
          info.textContent = searchResults.length ? `${{currentResultIndex + 1}}/${{searchResults.length}}` : '无匹配项';
          document.getElementById('prevButton').disabled = searchResults.length === 0 || currentResultIndex <= 0;
          document.getElementById('nextButton').disabled = searchResults.length === 0 || currentResultIndex >= searchResults.length - 1;
        }}
        function nextResult() {{
          if (searchResults.length === 0 || currentResultIndex >= searchResults.length - 1) return;
          currentResultIndex++; updateCurrentHighlight(); searchResults[currentResultIndex].element.scrollIntoView({{ behavior: 'smooth', block: 'center' }}); updateSearchInfo();
        }}
        function prevResult() {{
          if (searchResults.length === 0 || currentResultIndex <= 0) return;
          currentResultIndex--; updateCurrentHighlight(); searchResults[currentResultIndex].element.scrollIntoView({{ behavior: 'smooth', block: 'center' }}); updateSearchInfo();
        }}
        document.addEventListener('DOMContentLoaded', () => {{
          const s = document.getElementById('searchInput'); if (s) s.addEventListener('keypress', e => {{ if (e.key === 'Enter') performSearch(); }});
          document.addEventListener('keydown', e => {{ if (e.key === 'Escape') clearCrossHighlight(); }});
        }});
        </script>
        """)

    html = f"""<!DOCTYPE html><html lang="zh-CN"><head><meta charset="UTF-8">
        <title>企划数据看板</title><meta name="viewport" content="width=device-width,initial-scale=1">
        <style>{css}</style></head><body><div class="container">
        <div class="header"><h1>企划数据看板</h1>
          <div id="searchControls" style="display:none;flex:1 1 1;justify-content:center;"><div class="search-container">
            <input type="text" id="searchInput" class="search-input" placeholder="输入搜索关键词">
            <button class="search-button" onclick="performSearch()">搜索</button>
            <div class="nav-buttons">
              <button id="prevButton" class="nav-button" onclick="prevResult()" disabled>上一个</button>
              <div id="searchInfo" class="search-info">无匹配项</div>
              <button id="nextButton" class="nav-button" onclick="nextResult()" disabled>下一个</button>
            </div>
            <button class="search-button" onclick="clearSearch()" style="background:#dc3545;">清除</button>
          </div></div>
          <button id="backButton" class="back-button" onclick="goBack()" style="display:none;"><span>←</span><span>返回首页</span></button>
          <div class="header-time">{datetime.datetime.now():%Y-%m-%d %H:%M:%S}</div>
        </div>
        <div class="grid-wrapper"><div class="grid">
        <div style="margin:1rem 0;text-align:center;"><button class="view-button" style="background:#2db1d2">增删改后需等待2到3分钟</button></div>
        <div style="margin:1rem 0;text-align:center;"><button class="view-button" style="background:#28a745" onclick="uploadFile()">📤 上传新Excel</button></div>
        <div style="margin:1rem 0;text-align:center;"><button class="view-button" style="background:#6c757d" onclick="clearSavedToken()">🗑️ 清除保存的Token</button></div>
        {cards_html}
        </div></div>
        <div id="tableContainer" class="table-container"></div>
        <div class="footer">总计 {len(files)} 个Excel文件 | 数据看板系统</div>
        </div>{js}</body></html>"""

    with open(OUT_FILE, "w", encoding="utf-8") as f:
        f.write(html)
    print("✅ ExcelAPI.html 已生成（含完整 JS，无占位注释）")


if __name__ == "__main__":
    build_dashboard()